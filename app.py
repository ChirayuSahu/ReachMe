from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import os



app = Flask(__name__)
app.secret_key = "supersecretkey"

# Paths to Excel files
MEDICINE_PRICES_FILE = "medicine_prices.xlsx"
ORDERS_FILE = "orders.xlsx"
DOCTORS_LIST_FILE = "doctor_list.xlsx"

# Function to load medicine prices from Excel
def load_medicine_prices():
    if not os.path.exists(MEDICINE_PRICES_FILE):
        return {}

    workbook = openpyxl.load_workbook(MEDICINE_PRICES_FILE)
    sheet = workbook.active
    medicine_prices = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        medicine_name = row[0]
        price = row[1]
        medicine_prices[medicine_name.lower()] = price
    return medicine_prices

# Function to log the order to orders.xlsx
def log_order_to_excel(full_name, phone_number, medication, quantity, shipping_address):
    if not os.path.exists(ORDERS_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Full Name", "Phone Number", "Medication", "Quantity", "Shipping Address"])
        workbook.save(ORDERS_FILE)

    workbook = openpyxl.load_workbook(ORDERS_FILE)
    sheet = workbook.active
    sheet.append([full_name, phone_number, medication, quantity, shipping_address])
    workbook.save(ORDERS_FILE)

# Function to read doctor list from Excel using openpyxl
def get_doctor_list():
    # Load the workbook and select the active worksheet
    wb = load_workbook('doctor_list.xlsx')
    ws = wb.active

    # Collect doctor name and specialization from each row
    doctor_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Assuming the first row is the header
        doctor_name = row[0]
        specialization = row[1]
        doctor_list.append((doctor_name, specialization))

    return doctor_list

# Load or create the Excel file
excel_file = 'consultations.xlsx'

def init_excel():
    if not os.path.exists(excel_file):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Consultations'
        sheet.append(['Name', 'Phone', 'Problem', 'Doctor', 'Date', 'Time Slot'])
        workbook.save(excel_file)


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/orderMeds')
def order_meds():
    return render_template('orderMeds.html')

@app.route('/contact-us')
def contact_us():
    return render_template('contact-us.html')

@app.route('/about-us')
def about_us():
    return render_template('about-us.html')

@app.route('/getting-started')
def talk_ai():
    return render_template('getting-started.html')

@app.route('/consult')
def consult():
    # Load doctor data from Excel
    workbook = load_workbook('doctor_list.xlsx')
    sheet = workbook.active
    doctors = []
    
    for row in sheet.iter_rows(min_row=2, values_only=True):  # assuming the first row is headers
        doctor = {
            'name': row[0],        # Adjust indices based on your Excel structure
            'specialty': row[1]
        }
        doctors.append(doctor)
    return render_template('consult.html',doctors=doctors)

@app.route('/submit-consult', methods=['POST'])
def submit_consult():
    # Get form data
    name = request.form['name']
    phone = request.form['phone']
    problem = request.form['problem']
    doctor = request.form['doctor']
    date = request.form['date']
    time_slot = request.form['time_slot']
    
    # Store data in Excel
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    sheet.append([name, phone, problem, doctor, date, time_slot])
    workbook.save(excel_file)
    
    # Redirect to thank you page
    return redirect(url_for('thankyou', name=name, doctor=doctor, date=date, time_slot=time_slot))

@app.route('/thankyou')
def thankyou():
    # Get the data from the query string
    name = request.args.get('name')
    doctor = request.args.get('doctor')
    date = request.args.get('date')
    time_slot = request.args.get('time_slot')
    return render_template('thankyou.html', name=name, doctor=doctor, date=date, time_slot=time_slot)




@app.route('/submit', methods=['POST'])
def submit():
    full_name = request.form.get('Full_name')
    phone_number = request.form.get('Phone_number')
    shipping_address = request.form.get('Shipping_address')

    # Load medicine prices from the Excel file
    medicine_prices = load_medicine_prices()
    
    prescriptions = []
    
    # Loop through submitted medications
    medications = request.form.getlist('Medication')
    quantities = request.form.getlist('Quantity')

    for medication, quantity in zip(medications, quantities):
        medication = medication.lower()
        quantity = int(quantity)

        if medication in medicine_prices:
            price = medicine_prices[medication] * quantity
            prescription = f"Medicine: {medication.capitalize()}, Quantity: {quantity}, Total Price: ₹{price}"
            prescriptions.append(prescription)

            log_order_to_excel(full_name, phone_number, medication, quantity, shipping_address)
        else:
            flash(f"Sorry, {medication.capitalize()} is currently unavailable.", 'error')  # Flash error message
            return redirect(url_for('order_meds'))

    return render_template('prescription.html', prescriptions=prescriptions)

if __name__ == '__main__':
    app.run(debug=True)