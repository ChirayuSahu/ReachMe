from flask import Flask, render_template, request, redirect, url_for, flash
import openpyxl
import os

app = Flask(__name__)
app.secret_key = "supersecretkey"

# Paths to Excel files
MEDICINE_PRICES_FILE = "medicine_prices.xlsx"
ORDERS_FILE = "orders.xlsx"

# Function to load medicine prices from Excel
def load_medicine_prices():
    workbook = openpyxl.load_workbook(MEDICINE_PRICES_FILE)
    sheet = workbook.active
    medicine_prices = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        medicine_name = row[0]
        price = row[1]
        medicine_prices[medicine_name.lower()] = price  # Convert to lowercase for case-insensitive search
    return medicine_prices

# Function to log the order to orders.xlsx
def log_order_to_excel(full_name, phone_number, medication, shipping_address):
    # Check if the file exists, if not, create it
    if not os.path.exists(ORDERS_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Full Name", "Phone Number", "Medication", "Shipping Address"])  # Header
        workbook.save(ORDERS_FILE)
    
    # Load the orders file and append the new order
    workbook = openpyxl.load_workbook(ORDERS_FILE)
    sheet = workbook.active
    sheet.append([full_name, phone_number, medication, shipping_address])
    workbook.save(ORDERS_FILE)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    full_name = request.form.get('Full_name')
    phone_number = request.form.get('Phone_number')
    shipping_address = request.form.get('Shipping_address')

    # Load medicine prices from the Excel file
    medicine_prices = load_medicine_prices()
    
    prescriptions = []
    
    # Loop through submitted medications
    medications = request.form.getlist('Medication')
    quantities = request.form.getlist('Quantity')

    for medication, quantity in zip(medications, quantities):
        medication = medication.lower()  # Convert to lowercase for case-insensitive search
        quantity = int(quantity)  # Convert quantity to integer

        # Check if the entered medicine exists in the Excel list
        if medication in medicine_prices:
            price = medicine_prices[medication] * quantity
            prescription = f"Prescription for {full_name}: \nMedicine: {medication.capitalize()} \nQuantity: {quantity} \nTotal Price: ₹{price}"
            prescriptions.append(prescription)
            
            # Log the order to orders.xlsx
            log_order_to_excel(full_name, phone_number, medication, shipping_address)
        else:
            flash(f"Sorry, {medication.capitalize()} is currently unavailable.")
            return redirect(url_for('index'))

    # Show all prescriptions
    return render_template('prescription.html', prescriptions=prescriptions)

if __name__ == '__main__':
    app.run(debug=True)
